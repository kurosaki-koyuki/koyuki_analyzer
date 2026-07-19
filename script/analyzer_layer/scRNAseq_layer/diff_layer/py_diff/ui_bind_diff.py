# -*- coding: utf-8 -*-
"""
差异分析界面功能绑定脚本 - 全权负责粘合内外
绑定信号 + 编排 analysis 与 func 的协作
"""

from script.utils_layer.import_config import *
from script.mods_layer.mod_manager import global_mod_manager
from script.analyzer_layer.scRNAseq_layer.diff_layer.py_diff.diff_analysis import DiffAnalysis
from script.analyzer_layer.scRNAseq_layer.diff_layer.py_diff.ui_func_diff import DiffFunc
from script.utils_layer.music_controller_fix import fix_music_controller_bindings
from script.utils_layer.gui_styles import bind_button_with_sound
from script.utils_layer.page_intersect import page_intersect


class DiffBind:
    """差异分析功能绑定类 - 全权负责粘合内外"""

    def __init__(self, parent_window, diff_ui):
        self.parent = parent_window
        self.diff_ui = diff_ui
        self.analysis = DiffAnalysis()
        self.func = DiffFunc(diff_ui, parent_window)
        self.init_bindings()

    def init_bindings(self):
        """初始化所有绑定"""
        self.bind_music_controls()
        self.bind_diff_functions()
        self.bind_navigation()

    def bind_navigation(self):
        """绑定页面导航按钮"""
        if hasattr(self.diff_ui, 'nav_btn_back'):
            self.diff_ui.nav_btn_back.clicked.connect(lambda: page_intersect.go_to_page_with_bind('scRNAseq_top_page'))
        
        if hasattr(self.diff_ui, 'nav_btn_python'):
            self.diff_ui.nav_btn_python.clicked.connect(self.on_python_tab_click)
        
        if hasattr(self.diff_ui, 'nav_btn_r'):
            self.diff_ui.nav_btn_r.clicked.connect(self.on_r_tab_click)
    
    def on_python_tab_click(self):
        """Python版本标签点击处理"""
        if hasattr(self.diff_ui, 'nav_btn_python'):
            self.diff_ui.nav_btn_python.setChecked(True)
    
    def on_r_tab_click(self):
        """R版本标签点击处理"""
        page_intersect.go_to_page_with_bind('scRNAseq_r_diff_page')

    def sync_data_from_single_cell_main(self, single_cell_bind=None):
        """从单细胞分析主界面同步数据"""
        try:
            if single_cell_bind is None:
                single_cell_bind = getattr(self.parent, 'scRNAseq_top_bind', None)
            
            if single_cell_bind is None:
                return

            if hasattr(single_cell_bind, 'analysis') and single_cell_bind.analysis.adata is not None:
                self.analysis.set_adata(single_cell_bind.analysis.adata)
                self.load_groups()
                self.load_filter_columns()
            
            if hasattr(single_cell_bind, 'analysis') and single_cell_bind.analysis.dataset_output_dir is not None:
                self.analysis.set_dataset_output_dir(single_cell_bind.analysis.dataset_output_dir)

        except Exception as e:
            print(f"差异分析同步数据时出错: {str(e)}")

    def sync_data_from_analysis(self, analysis_bind=None):
        """从初步分析页面同步数据（兼容旧接口）"""
        self.sync_data_from_single_cell_main(analysis_bind)

    def bind_music_controls(self):
        """绑定音乐控制"""
        if hasattr(self.diff_ui, 'music_controller'):
            fix_music_controller_bindings(self, self.diff_ui.music_controller)

    def bind_diff_functions(self):
        """绑定差异分析功能信号"""
        log_widget = getattr(self.diff_ui, 'diff_log', None)

        if hasattr(self.diff_ui, 'btn_run_diff'):
            bind_button_with_sound(self.diff_ui.btn_run_diff, self.run_diff_analysis,
                                   log_widget, "差异分析完成", "差异分析失败")

        if hasattr(self.diff_ui, 'btn_export_csv'):
            bind_button_with_sound(self.diff_ui.btn_export_csv, self.export_results,
                                   log_widget, "导出完成", "导出失败")

        if hasattr(self.diff_ui, 'btn_export_png'):
            bind_button_with_sound(self.diff_ui.btn_export_png, self.export_png,
                                   log_widget, "PNG导出完成", "PNG导出失败")

        if hasattr(self.diff_ui, 'diff_group_combo'):
            self.diff_ui.diff_group_combo.currentIndexChanged.connect(self.on_group_combo_changed)
        
        if hasattr(self.diff_ui, 'diff_filter1_col'):
            self.diff_ui.diff_filter1_col.currentIndexChanged.connect(self.on_filter1_col_changed)
        
        if hasattr(self.diff_ui, 'diff_filter2_col'):
            self.diff_ui.diff_filter2_col.currentIndexChanged.connect(self.on_filter2_col_changed)

        if hasattr(self.diff_ui, 'gene_search_btn') and self.func:
            self.diff_ui.gene_search_btn.clicked.connect(self.func.search_gene)

        if hasattr(self.diff_ui, 'gene_search_input') and self.func:
            self.diff_ui.gene_search_input.returnPressed.connect(self.func.search_gene)

    def on_filter1_col_changed(self):
        """筛选条件1列变化时更新值列表"""
        self.load_filter_values(self.diff_ui.diff_filter1_col, self.diff_ui.diff_filter1_list)
    
    def on_filter2_col_changed(self):
        """筛选条件2列变化时更新值列表"""
        self.load_filter_values(self.diff_ui.diff_filter2_col, self.diff_ui.diff_filter2_list)
    
    def load_filter_columns(self):
        """加载可用于筛选的列"""
        try:
            if self.analysis.adata is None:
                return
            
            filter_cols = [''] + self.analysis.get_available_groups()
            
            self.func.set_combo_items(self.diff_ui.diff_filter1_col, filter_cols)
            self.func.set_combo_items(self.diff_ui.diff_filter2_col, filter_cols)
            
            self.load_filter_values(self.diff_ui.diff_filter1_col, self.diff_ui.diff_filter1_list)
            self.load_filter_values(self.diff_ui.diff_filter2_col, self.diff_ui.diff_filter2_list)
            
        except Exception as e:
            self.func.log(f"加载筛选列失败: {str(e)}")
    
    def load_filter_values(self, col_combo, list_widget):
        """加载指定筛选列的唯一值到多选列表"""
        try:
            col_name = col_combo.currentText()
            if not col_name or self.analysis.adata is None:
                self.func.fill_list_widget(list_widget, [], select_all=False)
                return
            
            unique_vals = self.analysis.get_group_unique_vals(col_name)
            self.func.fill_list_widget(list_widget, unique_vals, select_all=False)
            
        except Exception as e:
            self.func.log(f"加载筛选值失败: {str(e)}")

    def on_group_combo_changed(self):
        """分组下拉框变化时更新分组列表"""
        self.load_groups()

    def load_groups(self):
        """加载分组列表"""
        try:
            if self.analysis.adata is None:
                return

            groups = self.analysis.get_available_groups()
            if not groups:
                return

            self.func.set_combo_items(self.diff_ui.diff_group_combo, groups)

            current_group = self.diff_ui.diff_group_combo.currentText()
            if current_group:
                unique_vals = self.analysis.get_group_unique_vals(current_group)
                self.func.update_group_lists(current_group, unique_vals)

            self.func.log(f"加载分组完成，共 {len(groups)} 个可选列")

        except Exception as e:
            self.func.alert_failure(f"加载分组失败: {str(e)}")
            self.func.log(f"❌ {str(e)}")

    def get_filter_mask(self):
        """获取细胞筛选掩码（支持多选）"""
        mask = pd.Series([True] * len(self.analysis.adata), index=self.analysis.adata.obs.index)
        
        filter1_col = self.diff_ui.diff_filter1_col.currentText()
        filter1_vals = [item.text() for item in self.diff_ui.diff_filter1_list.selectedItems()]
        if filter1_col and filter1_vals:
            mask = mask & self.analysis.adata.obs[filter1_col].isin(filter1_vals)
        
        filter2_col = self.diff_ui.diff_filter2_col.currentText()
        filter2_vals = [item.text() for item in self.diff_ui.diff_filter2_list.selectedItems()]
        if filter2_col and filter2_vals:
            mask = mask & self.analysis.adata.obs[filter2_col].isin(filter2_vals)
        
        return mask

    def run_diff_analysis(self):
        """执行差异分析"""
        try:
            group_col = self.diff_ui.diff_group_combo.currentText()
            
            group1_items = [item.text() for item in self.diff_ui.diff_group1_list.selectedItems()]
            group2_items = [item.text() for item in self.diff_ui.diff_group2_list.selectedItems()]

            if not group_col:
                self.func.alert_error("请先选择分组列")
                return

            if len(group1_items) == 0:
                self.func.alert_error("请至少选择1个组别1")
                return

            if len(group2_items) == 0:
                self.func.alert_error("请至少选择1个组别2")
                return

            method = "mannwhitney"
            min_cells = self.diff_ui.diff_min_cells.value()
            min_expr = self.diff_ui.diff_min_expr.value()
            use_fdr = self.diff_ui.diff_use_fdr.isChecked()
            
            pval_threshold = self.diff_ui.diff_pval_spin.value()
            logfc_threshold = self.diff_ui.diff_logfc_spin.value()
            pct_threshold = self.diff_ui.diff_pct_spin.value()

            method_display = "Mann-Whitney U检验" if method == "mannwhitney" else "scanpy-mast检验"
            self.func.log("开始差异分析...")
            self.func.log(f"分组: {group_col}")
            self.func.log(f"组别1: {', '.join(group1_items)}")
            self.func.log(f"组别2: {', '.join(group2_items)}")
            self.func.log(f"方法: {method_display}, min_cells={min_cells}")
            self.func.log(f"阈值: pval={pval_threshold}, logfc={logfc_threshold}, pct={pct_threshold}")

            filter_mask = self.get_filter_mask()

            results_df = self.analysis.run_diff_analysis(
                group_col=group_col,
                selected_groups=[group1_items, group2_items],
                method=method,
                min_cells=min_cells,
                min_expr=min_expr,
                use_fdr=use_fdr,
                pval_threshold=pval_threshold,
                logfc_threshold=logfc_threshold,
                pct_threshold=pct_threshold,
                filter_mask=filter_mask
            )

            self.func.update_diff_stats(results_df, group1_items, group2_items)
            self.func.fill_result_tables(
                results_df,
                results_df[results_df['change'] == 'up'],
                results_df[results_df['change'] == 'down']
            )
            self.func.render_volcano_plot(results_df, group_col, [group1_items, group2_items])

            self.func.log(f"找到 {len(results_df)} 个差异基因")

        except ValueError as e:
            self.func.alert_error(str(e))
            self.func.log(f"❌ ValueError: {str(e)}")
        except Exception as e:
            import traceback
            self.func.alert_failure(f"差异分析失败: {str(e)}")
            self.func.log(f"❌ {str(e)}")
            self.func.log(f"详细错误:\n{traceback.format_exc()}")

    def export_results(self):
        """导出差异分析结果到xlsx文件（包含统计信息、总体列表、上调、下调）"""
        if self.analysis.diff_gene_df is None or len(self.analysis.diff_gene_df) == 0:
            self.func.alert_error("请先执行差异分析")
            return

        try:
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            self.func.alert_error("请安装 openpyxl 库以导出xlsx文件")
            return

        group1_items = [item.text() for item in self.diff_ui.diff_group1_list.selectedItems()]
        group2_items = [item.text() for item in self.diff_ui.diff_group2_list.selectedItems()]
        
        def build_group_name(categories, default_name):
            if not categories:
                return default_name
            if len(categories) == 1:
                return categories[0]
            full_name = "+".join(categories)
            if len(full_name) <= 28:
                return full_name
            short_name_3 = "+".join([cat[:3] for cat in categories])
            if len(short_name_3) <= 28:
                return short_name_3
            return "+".join([cat[:1] for cat in categories])
        
        group1_name = build_group_name(group1_items, "组1")
        group2_name = build_group_name(group2_items, "组2")

        filter_info = []
        filter1_col = self.diff_ui.diff_filter1_col.currentText()
        filter1_vals = [item.text() for item in self.diff_ui.diff_filter1_list.selectedItems()]
        if filter1_col and filter1_vals:
            filter_info.append(f"{filter1_col}={','.join(filter1_vals)}")
        
        filter2_col = self.diff_ui.diff_filter2_col.currentText()
        filter2_vals = [item.text() for item in self.diff_ui.diff_filter2_list.selectedItems()]
        if filter2_col and filter2_vals:
            filter_info.append(f"{filter2_col}={','.join(filter2_vals)}")
        
        filter_str = "_筛选=" + ";".join(filter_info) if filter_info else "_筛选=无"
        default_name = f"差异分析_{group1_name}vs{group2_name}{filter_str}.xlsx"

        save_path = self.func.get_save_file_path("导出差异分析结果", default_name, "Excel文件 (*.xlsx)")

        if save_path:
            try:
                wb = openpyxl.Workbook()
                
                ws1 = wb.active
                ws1.title = "统计信息"
                stats_data = [
                    ["统计项", "数值"],
                    ["比较组1", group1_name],
                    ["比较组2", group2_name],
                    [self.diff_ui.diff_group1_cell_label.text().split(":")[0], 
                     self.diff_ui.diff_group1_cell_label.text().split(":")[1].strip()],
                    [self.diff_ui.diff_group2_cell_label.text().split(":")[0], 
                     self.diff_ui.diff_group2_cell_label.text().split(":")[1].strip()],
                    [self.diff_ui.diff_up_label.text().split(":")[0], 
                     self.diff_ui.diff_up_label.text().split(":")[1].strip()],
                    [self.diff_ui.diff_down_label.text().split(":")[0], 
                     self.diff_ui.diff_down_label.text().split(":")[1].strip()],
                    [self.diff_ui.diff_stable_label.text().split(":")[0], 
                     self.diff_ui.diff_stable_label.text().split(":")[1].strip()],
                    [self.diff_ui.diff_total_label.text().split(":")[0], 
                     self.diff_ui.diff_total_label.text().split(":")[1].strip()],
                    ["筛选条件", ";".join(filter_info) if filter_info else "无"]
                ]
                for row in stats_data:
                    ws1.append(row)
                
                ws2 = wb.create_sheet(title="总体差异分析列表")
                for r in dataframe_to_rows(self.analysis.diff_gene_df, index=False, header=True):
                    ws2.append(r)
                
                ws3 = wb.create_sheet(title=f"{group1_name}显著上调基因")
                up_df = self.analysis.diff_gene_df[self.analysis.diff_gene_df['change'] == 'up']
                for r in dataframe_to_rows(up_df, index=False, header=True):
                    ws3.append(r)
                
                ws4 = wb.create_sheet(title=f"{group1_name}显著下调基因")
                down_df = self.analysis.diff_gene_df[self.analysis.diff_gene_df['change'] == 'down']
                for r in dataframe_to_rows(down_df, index=False, header=True):
                    ws4.append(r)
                
                wb.save(save_path)
                self.func.alert_success(f"结果已保存到:\n{save_path}")
            except Exception as e:
                self.func.alert_failure(f"导出失败: {str(e)}")

    def export_png(self):
        """导出火山图为PNG"""
        if self.analysis.diff_gene_df is None or len(self.analysis.diff_gene_df) == 0:
            self.func.alert_error("请先执行差异分析")
            return

        group1_items = [item.text() for item in self.diff_ui.diff_group1_list.selectedItems()]
        group2_items = [item.text() for item in self.diff_ui.diff_group2_list.selectedItems()]
        
        def build_group_name(categories, default_name):
            if not categories:
                return default_name
            if len(categories) == 1:
                return categories[0]
            full_name = "+".join(categories)
            if len(full_name) <= 28:
                return full_name
            short_name_3 = "+".join([cat[:3] for cat in categories])
            if len(short_name_3) <= 28:
                return short_name_3
            return "+".join([cat[:1] for cat in categories])
        
        group1_name = build_group_name(group1_items, "组1")
        group2_name = build_group_name(group2_items, "组2")
        
        filter_info = []
        filter1_col = self.diff_ui.diff_filter1_col.currentText()
        filter1_vals = [item.text() for item in self.diff_ui.diff_filter1_list.selectedItems()]
        if filter1_col and filter1_vals:
            filter_info.append(f"{filter1_col}={','.join(filter1_vals)}")
        
        filter2_col = self.diff_ui.diff_filter2_col.currentText()
        filter2_vals = [item.text() for item in self.diff_ui.diff_filter2_list.selectedItems()]
        if filter2_col and filter2_vals:
            filter_info.append(f"{filter2_col}={','.join(filter2_vals)}")
        
        filter_str = "_筛选=" + ";".join(filter_info) if filter_info else "_筛选=无"
        default_name = f"火山图_{group1_name}vs{group2_name}{filter_str}.png"
        save_path = self.func.get_save_file_path("导出火山图", default_name, "PNG文件 (*.png)")

        if save_path:
            try:
                self.analysis.export_png(save_path)
                self.func.alert_success(f"火山图已保存到:\n{save_path}")
            except Exception as e:
                self.func.alert_failure(f"导出失败: {str(e)}")

    def set_volume(self, value):
        """设置音量"""
        mod_instance = global_mod_manager.get_current_mod()
        if hasattr(mod_instance, 'global_music_player'):
            mod_instance.global_music_player.set_volume(value / 100.0)

        if hasattr(self.parent, '_sync_all_volume_sliders_from_subinterface'):
            self.parent._sync_all_volume_sliders_from_subinterface(value)

    def set_adata(self, adata):
        """设置adata对象"""
        self.analysis.set_adata(adata)

    def set_dataset_output_dir(self, dataset_output_dir):
        """设置数据集输出目录"""
        self.analysis.set_dataset_output_dir(dataset_output_dir)
